"""Excel report renderer."""

import datetime
import os
from typing import List

from ..config import CONFIG
from ..models import HostGroup, PackageGroup, ScoredVuln
from .csv_output import CSV_HEADERS

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    pass  # Checked at startup via auth.check_required_tools


def write_excel(
    vulns: List[ScoredVuln],
    pkg_groups: List[PackageGroup],
    host_groups: List[HostGroup],
    output_path: str,
    min_severity: str,
) -> None:
    """Write 4-sheet Excel workbook: Summary, By Package, By Host, Detail."""
    wb = Workbook()

    header_fill = PatternFill(
        start_color=CONFIG.EXCEL_HEADER_COLOR,
        end_color=CONFIG.EXCEL_HEADER_COLOR,
        fill_type="solid",
    )
    header_font = Font(bold=True, color=CONFIG.EXCEL_HEADER_TEXT_COLOR)
    link_font = Font(color=CONFIG.EXCEL_LINK_COLOR, underline="single")

    def write_headers(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

    def adjust_widths(ws, num_cols):
        for col in range(1, num_cols + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, CONFIG.EXCEL_MAX_COLUMN_WIDTH)

    def add_autofilter(ws, num_cols):
        last_col = get_column_letter(num_cols)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = ["Metric", "Value"]
    write_headers(ws_summary, summary_headers)

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    total_cves = len(set(v.vuln_id for v in vulns))
    total_pkgs = len(set((v.pkg_name, v.pkg_namespace) for v in vulns))
    total_hosts = len(set(v.machine_id for v in vulns))
    critical = sum(1 for v in vulns if v.severity == "Critical")
    high = sum(1 for v in vulns if v.severity == "High")

    summary_data = [
        ("Report Generated", now),
        ("Minimum Severity", min_severity),
        ("Fixable Only", "Yes"),
        ("Total Entries", len(vulns)),
        ("Unique CVEs", total_cves),
        ("Vulnerable Packages", total_pkgs),
        ("Affected Hosts", total_hosts),
        ("Critical Findings", critical),
        ("High Findings", high),
    ]
    for row, (metric, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row, column=1, value=metric)
        ws_summary.cell(row=row, column=2, value=value)
    adjust_widths(ws_summary, 2)

    # --- Sheet 2: By Package ---
    ws_pkg = wb.create_sheet("By Package")
    pkg_headers = [
        "Rank", "Package", "Namespace", "Fix Version", "CVEs",
        "Hosts", "Max CVSS", "Aggregate Score",
    ]
    write_headers(ws_pkg, pkg_headers)

    for row, g in enumerate(pkg_groups, 2):
        ws_pkg.cell(row=row, column=1, value=row - 1)
        ws_pkg.cell(row=row, column=2, value=g.pkg_name)
        ws_pkg.cell(row=row, column=3, value=g.pkg_namespace)
        ws_pkg.cell(row=row, column=4, value=g.fixed_version)
        ws_pkg.cell(row=row, column=5, value=", ".join(sorted(g.cve_ids)))
        ws_pkg.cell(row=row, column=6, value=len(g.hosts))
        ws_pkg.cell(row=row, column=7, value=g.max_cvss)
        ws_pkg.cell(row=row, column=8, value=g.aggregate_score)

    adjust_widths(ws_pkg, len(pkg_headers))
    add_autofilter(ws_pkg, len(pkg_headers))

    # --- Sheet 3: By Host ---
    ws_host = wb.create_sheet("By Host")
    host_headers = [
        "Rank", "Hostname", "Instance ID",
        "VM Provider", "Collector", "Unique CVEs", "Unique Packages",
        "Host Risk Score", "Max Priority",
    ]
    write_headers(ws_host, host_headers)

    for row, h in enumerate(host_groups, 2):
        ws_host.cell(row=row, column=1, value=row - 1)
        ws_host.cell(row=row, column=2, value=h.hostname)
        ws_host.cell(row=row, column=3, value=h.instance_id)
        ws_host.cell(row=row, column=4, value=h.vm_provider)
        ws_host.cell(row=row, column=5, value=h.collector_type)
        ws_host.cell(row=row, column=6, value=h.unique_cves)
        ws_host.cell(row=row, column=7, value=h.unique_packages)
        ws_host.cell(row=row, column=8, value=h.host_risk_score)
        ws_host.cell(row=row, column=9, value=h.max_priority)

    adjust_widths(ws_host, len(host_headers))
    add_autofilter(ws_host, len(host_headers))

    # --- Sheet 4: Detail (flat) ---
    ws_detail = wb.create_sheet("Detail")
    write_headers(ws_detail, CSV_HEADERS)

    sorted_vulns = sorted(vulns, key=lambda v: v.priority_score, reverse=True)
    for row, v in enumerate(sorted_vulns, 2):
        ws_detail.cell(row=row, column=1, value=v.priority_score)
        ws_detail.cell(row=row, column=2, value=v.vuln_id)
        ws_detail.cell(row=row, column=3, value=v.severity)
        ws_detail.cell(row=row, column=4, value=v.cvss_score)
        ws_detail.cell(row=row, column=5, value=v.pkg_name)
        ws_detail.cell(row=row, column=6, value=v.pkg_namespace)
        ws_detail.cell(row=row, column=7, value=v.pkg_version_installed)
        ws_detail.cell(row=row, column=8, value=v.fixed_version)
        ws_detail.cell(row=row, column=9, value=v.hostname)
        ws_detail.cell(row=row, column=10, value=v.instance_id)
        ws_detail.cell(row=row, column=11, value=v.first_seen)
        ws_detail.cell(row=row, column=12, value="Yes" if v.exploit_public else "No")
        ws_detail.cell(row=row, column=13, value="Yes" if v.exploit_wormified else "No")
        ws_detail.cell(row=row, column=14, value=v.vm_provider)
        ws_detail.cell(row=row, column=15, value=v.collector_type)

        if v.cve_link:
            cell = ws_detail.cell(row=row, column=16, value=v.cve_link)
            cell.hyperlink = v.cve_link
            cell.font = link_font
        else:
            ws_detail.cell(row=row, column=16, value="")

    adjust_widths(ws_detail, len(CSV_HEADERS))
    add_autofilter(ws_detail, len(CSV_HEADERS))

    # Save
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
